# -*- coding: utf-8 -*-
"""
Created on Wed Jun 19 14:35:01 2019

@author: Pedro
"""

import serial
import pandas as pd
import numpy as np
import time
from openpyxl import Workbook
import winsound
import statistics
import matplotlib.pyplot as plt    
        
#######################################################
#these functions communicate with the lock-in
#####################################################################

def ask_lockin(cmd):
    lock.write(str.encode(cmd+'\n'))  
    lock.flush()
    state = lock.readline()  
    decoded = bytes.decode(state)
    return(decoded)

def tell_lockin(cmd):
    lock.write(str.encode(cmd+'\n'))  
    lock.flush() 

#####################################################################
#these functions read and write to excel 
#####################################################################

def createtitles1(a): #data.append([val_f,j,out_x,sd_x,out_y,sd_y,sensitivity])
    b=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V3W Measurement"
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W Measurement"   
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W_shunt Measurement"    
    a+=1
    print("wrote titles")
    return a

def createtitles2(a):
    b=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V3W Measurement"
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W Measurement"   
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W_shunt Measurement"    
    b+=1
    sheet.cell(row=a, column=b).value = "T of sample" 
    a+=1
    print("wrote titles")
    return a

def to_excel(measurements,c,d): #c=a d=b
    for a, i in enumerate(measurements):
        for b, elem in enumerate(i):
            sheet.cell(row=(a+c), column=(b+d)).value = elem
    book.save('hello.xlsx')
    return (c + len(measurements))

def setting():
    ask_lockin('RMOD?') #returns dynamic reserve mode (high reserve/normal/low noise)
    ask_lockin('OFLT?') #returns the time constant (10microsec-30kilosec)
    ask_lockin('OFSL?') #returns low pass filter slope (6/12/18/24 dB/oct) 
    ask_lockin('SYNC?') #returns synchronous filter (Off/On) 
    ask_lockin('IGND?') #returns input shield grounding (GROUND/FLOATING)
    ask_lockin('ICPL?') #returns input coupling (AC/DC)
    ask_lockin('ILIN?') #returns line notch filters (Out, Line in, 2xLine in, Both In)
##########################################################
#these functions are for scanning and returning data values
################################################################
'''
def stable(x_splits,stablity_cond): #change how stability is defined********************************W******
    x = x_splits 
    vg = int(len(x)/2)

    high_avg = x[:vg].mean()
    low_avg = x[vg:len(x)].mean()
    
    print('next 3')
    print('')
    print(x[:vg])
    print('')
    print(x[vg:len(x)])
    print('')
    
    z = abs(high_avg - low_avg)
    if z < stablity_cond:
        return True
    else:
        return False
  '''
def stable(x_splits,stablity_cond): #change how stability is defined********************************W******
    x = x_splits 
    vg = int(len(x)/2)
    print('next 3')
    print('')
    print(x[:vg])
    print('')
    print(x[vg:len(x)])
    print('')
    maxi = np.amax(x_splits)
    mini = np.amin(x_splits)
    z = abs(maxi - mini)
    if z < stablity_cond:
        print("now stable")
        return True
    else:
        return False
    
def calibrate(stb,max_stb_time,lc_list,lc_cmd):
    t = max_stb_time
    print('')
    print('calibrating...')
    tell_lockin('REST')
    
    tell_lockin(lc_cmd + str(lc_list[0]))
    time.sleep(6)
    tell_lockin('STRT')
    tell_lockin(lc_cmd + str(lc_list[1]))
    time.sleep(t)
    x = ask_lockin('TRCA?1,0,'+str(31*t))
    x = x[:len(x)-2] #change back to 2 if you get a syntax error
    x = x.split(',')
    x = [float(i) for i in x]
    x = np.array(x)
    spaced = t-3
    for i in range(spaced):
        avg = x[31*(i+1):31*(i+4)]
        if stable(avg,stb) == True:
            print('')
            print('stabilized after '+ str(i+1) + ' seconds')
            return i+4
            break
    stb = float(input('reconsider stability condition as system is not stable after max time allowed. Input new stability condition:  '))
    max_stb_time = int(input('reconsider max time allowed as system is not stable after previously inputed max time. Input new max calibration time:  '))
    calibrate(stb,max_stb_time,lc_list,lc_cmd)
  
def scan(harm,lc_list,volt_q,val_f,stab_cond): #char is the command for what is being measured, listchar is the list containing the values that are to be scanned, stable_time is the amount of time spent at each value in listchar, tt is the amount of time taken into account for mean.x and stdv.x
    tell_lockin('HARM'+str(harm))
    output_x = []                                       #volt_q=True if scanning volts false if scanning freq
    stdv_x = []
    data = []
    
    tt = 3 #time that is averaged over for mean
    time_spent = 6
    
    #delete both arguments in function later
    #stab_cond = float(input("input stability condition(V):    "))
    #time_spent = int(input("input total time waiting for calibration(s):   "))
    #relocate these statements to outside scan so it doesnt have to ask for each sweep
    
    
    if volt_q == 1:
        lc_list = listvolts
        lc_cmd = 'SLVL'
        na = 'V'
    else:
        lc_list = listfreq
        lc_cmd = 'FREQ'
        na = 'Hz'
    
    t = calibrate(stab_cond,time_spent,lc_list,lc_cmd) #if shit is hitting the fan change back to 16
    
    print('scanning whole range')
    print('')
    
    sensitivity = ask_lockin('SENS?')
    tell_lockin('REST') #clears and pauses buffer
    tell_lockin('STRT') #starts scan
    
    for i in lc_list:
        tell_lockin(lc_cmd+str(i))
        time.sleep(t)
        
    x = ask_lockin('TRCA?1,0,'+str(31*len(lc_list)*t)) #outputs buffer
    x = x[:len(x)-1]
    print('done scanning')
    y = ask_lockin('TRCA?2,0,'+str(31*len(lc_list)*t)) 
    y = y[:len(y)-1]
    
    for i,j in enumerate(lc_list):
        x_split = x.split(',')[((31*i*t)+(t-tt)*31):(31*(i+1)*t)]    #31 is number of measurements a second
        
        x_split = [float(i) for i in x_split]
        x_split = np.array(x_split)
        
        if stable(x_split,stab_cond) == False:
            print("was not stable at: " + str(j)+na)
        else: 
            print("was stable at: " + str(j)+na)
            
        y_split = y.split(',')[((31*i*t)+(t-tt)):(31*(i+1)*t)]
        y_split = [float(i) for i in x_split]
        y_split = np.array(y_split)
        
        out_x = x_split.mean()
        out_y = y_split.mean()
        print('')
        print("output mean:  "+str(out_x))
        print('')
        sd_x = x_split.std()
        sd_y = y_split.std()
        
        #out_x,out_y,sd_x,sd_y= averaged_data(x_split,y_split)
        output_x.append(out_x)
        stdv_x.append(sd_x)
        
        
        if volt_q != 1:
            huh = val_f
            val_f = j
            j = huh
            
        data.append([val_f,j,out_x,sd_x,out_y,sd_y,sensitivity])
  
     
    return data,output_x,stdv_x

########################################################@##
#THESE functions calculate and plot values#################################
###########################################################

def calculate_plot1(v1_sh,r_sh,v1_):
    i_current = []
    i_cubbed = []
    r = []
    power = []
    ctr = 0
    for i in range(len(v1_)):
        for j in range(len(v1_[0])):
            print(type(v1_sh))
            i_current.append(v1_sh[i][j]/r_sh)     #i = v1w / r shunt
            i_cubbed.append(i_current[ctr]**3)            #i^3 
            power.append(v1_[i][j]*i_current[ctr])        #power = v1w * current
            r.append(v1_[i][j]/i_current[ctr])            #r = v1w / i
            ctr+=1
    return i_current,i_cubbed,power,r

'''
fig = plt.figure()
        plt.errorbar(listfreq[1:],y1,fmt='o') 
        plt.errorbar(listfreqrev[1:],y2,fmt='o') 
        fig.suptitle('Time to stabilize at different frequencies ('+str(harm)+'\u03C9)', fontsize=20)
        plt.xlabel('Frequency', fontsize=16)
        plt.ylabel('Time to stabilize', fontsize=16)
        plt.legend(['forward','reverse'])
        fig.savefig('calibrate_freq'+str(ctr)+'at harm'+str(harm)+'.jpg')
        plt.show()
'''

def plot_ic_vs_v3(v3w,cub,stdv_v3w,num_voltages,freqs):
    fig = plt.figure() 
    num_sweeps = len(v3w)/num_voltages
    v3w =(np.array(v3w)).flatten()
    stdv_v3w = (np.array(stdv_v3w)).flatten()
    cub = np.array(cub)
    for i in range(num_sweeps):
        plt.errorbar(cub[(i*num_voltages):((1+i)*num_voltages)],v3w[(i*num_voltages):((1+i)*num_voltages)], stdv_v3w[(i*num_voltages):((1+i)*num_voltages)],fmt='o') #0:4 -> 4-8
    fig.suptitle('I^3 vs V3\u03C9)', fontsize=20)
    plt.xlabel('I^3 (A)', fontsize=16)
    plt.ylabel('V3\u03C9(V)', fontsize=16)
    plt.legend([str(i) for i in freqs]) 
    fig.savefig('I^3 vs V3W.jpg')
    plt.show()

'''
def plot_ic_vs_v3(v3w,cub,stdv_v3w,num_voltages,freqs):
    fig = plt.figure() 
    num_sweeps = len(freqs)
    v3w =(np.array(v3w)).flatten()
    stdv_v3w = (np.array(stdv_v3w)).flatten()
    cub = np.array(cub)
    for i in range(num_sweeps):
        plt.errorbar(cub[(i*num_voltages):((1+i)*num_voltages)],v3w[(i*num_voltages):((1+i)*num_voltages)], stdv_v3w[(i*num_voltages):((1+i)*num_voltages)],fmt='o') #0:4 -> 4-8
    fig.suptitle('I^3 vs V3\u03C9)', fontsize=20)
    plt.xlabel('I^3 (A)', fontsize=16)
    plt.ylabel('V3\u03C9(V)', fontsize=16)
    plt.legend([str(i) for i in freqs]) 
    fig.savefig('I^3 vs V3W.jpg')
    plt.show()   
'''    

def plot_r_vs_power(r,power,num_voltages,freqs):
    fig = plt.figure() 
    num_sweeps = len(r)/num_voltages
    for i in range(num_sweeps):
        plt.errorbar(power[(i*num_voltages):((1+i)*num_voltages)],r[(i*num_voltages):((1+i)*num_voltages)],fmt='o') #0:4 -> 4-8
    fig.suptitle('Resistance of Sample vs Power)', fontsize=20)
    plt.xlabel('R Sample[Ohm]')
    plt.ylabel('Power[W]')
    plt.legend([str(i) for i in freqs]) 
    fig.savefig('R vs P.jpg')
    plt.show()
    
def plot_v3w_vs_ln2w(v3w,stdv_v3w,freqs,num_voltages):
    v3w = np.array(v3w[len(freqs)])
    stdv_v3w = np.array(stdv_v3w)
    ln2w = [np.log(4*i*np.pi) for i in freqs]
    ln2w = ln2w + (list(reversed(ln2w))) +  ln2w + (list(reversed(ln2w)))
    plt.errorbar(v3w.flatten(),ln2w,stdv_v3w.flatten(),fmt='o') 
    plt.show()
 
#initialize variables
a = 1
b = 1
book = Workbook()
sheet = book.active    



# main
lock = serial.Serial('COM5', baudrate=19200, parity=serial.PARITY_NONE,
                         stopbits=serial.STOPBITS_ONE, bytesize=serial.EIGHTBITS, timeout=3)   

A = pd.read_excel (r'C:\Users\Pedro\Desktop\HelloWorld\SR830\sweep.xlsx',sheet_name='Sheet1')
listfreq=A['freq']
listvolts=A['volts']

tell_lockin("OUTX0")
tell_lockin('SRAT9')


a = createtitles1(a)
r_shunt = float(input("input r_shunt:  "))


i = int(input('How many frequencies do you wanna test?:    '))
freq_1 = [j for j in range(i)]

for j in range(i):
    freq_1[j] = input('Input the next frequency you want to test:    ')

v1 = []
std_v1 = []
v3 = []
std_v3 = []
v1_shunt = []
std_v1_shunt = []


#remove harm from scan
#set next settings asap
#change amount of time taken to wait at first interval

aa = a
input("adjust sensitivity and when ready for 3w measurement of sample press enter (usually 50mV)")
for j in range(i):
    tell_lockin('FREQ'+freq_1[j])
    tabledata,v3w,std_v3w = scan(3,listvolts,1,freq_1[j],249e-6)
    a = to_excel(tabledata,a,1)
    v3.append(v3w)
    std_v3.append(std_v3w)
a = aa
print('')
print('finished 3w')
book.save('hello.xlsx')

input("Adjust sensitivity and when ready for 1w measurement of sample varying voltage press enter")
for j in range(i):   
    tell_lockin('FREQ'+freq_1[j])
    tabledata,v1w,std_v1w = scan(1,listvolts,1,freq_1[j],0.00005)
    a = to_excel(tabledata,a,8)
    v1.append(v1w)
    std_v1.append(std_v1w)
a = aa
print('')
print('finished 1w')
book.save('hello.xlsx')

input("Swap bananas. Adjust sensitivity. When ready for 1w shunt measurement press enter")
for j in range(i):    
    tell_lockin('FREQ'+freq_1[j])
    tabledata,v1w_shunt,std_v1w_shunt = scan(1,listvolts,1,freq_1[j],0.00005)
    a = to_excel(tabledata,a,15)
    v1_shunt.append(v1w_shunt)
    std_v1_shunt.append(std_v1w_shunt)
print('')
print('finished 1w_shunt')
book.save('hello.xlsx')

current,cubbed,p,r = calculate_plot1(v1_shunt,r_shunt,v1)
plot_ic_vs_v3(v3,cubbed,std_v3,len(listvolts),len(listvolts),freq_1)
plot_r_vs_power(r,p,len(listvolts))


a = createtitles2(a)
aa = a
volt = input("what is the voltage you have chosen to work with?:  ")
tell_lockin("SLVL"+str(volt))
temp = input("input sample temperature:  ")
uio = to_excel(tabledata,a,19)
input("Adjust sensitivity and when ready for 3w sample measurement press enter")
book.save('hello.xlsx')

v3_sample=[]
std_v3=[]
v3_sample_sweep=[]
std_v3_sweep=[]
for i in range(4):
    tabledata,v3_sample,std_v3 = scan(3,listfreq,0,volt,0.002e-3)     
    a = to_excel(tabledata,a,1)
    listfreq = list(reversed(listfreq))
    v3_sample_sweep.append(v3_sample)
    std_v3_sweep.append(std_v3)
a = aa
print('done 3w sample measurement')
book.save('hello.xlsx')
###############################################

input("Adjust sensitivity and when ready for 1w sample measurement press enter")
tabledata,v1w_sample,std_v1w_samplze = scan(1,listfreq,0,volt,0.0001)     
a = to_excel(tabledata,a,8)
a = aa
print('done 1w sample measurement')
book.save('hello.xlsx')
###############################################

input("Adjust sensitivity and when ready for 1w shunt measurement press enter. make sure bannas are in place")
tabledata,v1w_shunt,std_v1w_shunt = scan(1,listfreq,0,volt,0.0001)     
a = to_excel(tabledata,a,15)

print('done 1w shunt measurement')
book.save('hello.xlsx')

print(v3_sample_sweep)
print(std_v3_sweep)
plot_v3w_vs_ln2w(v3_sample_sweep,std_v3_sweep,listfreq)


book.save('hello.xlsx')
lock.close()