# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 09:04:49 2019

@author: Pedro
"""

import serial
import pandas as pd
import numpy as np
import time
from openpyxl import Workbook
import xlsxwriter 
import winsound
import statistics
import matplotlib.pyplot as plt    
        
#######################################################
#these functions communicate with the lock-in
#####################################################################

def ask_lockin(cmd):
    lock.write(str.encode(cmd+'\n'))  
    lock.flush()
    state = lock.readline()  
    decoded = bytes.decode(state)
    return(decoded)

def tell_lockin(cmd):
    lock.write(str.encode(cmd+'\n'))  
    lock.flush() 

#####################################################################
#these functions read and write to excel 
#####################################################################

def createtitles1(a): #data.append([val_f,j,out_x,sd_x,out_y,sd_y,sensitivity])
    b=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V3W Measurement"
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W Measurement"   
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W_shunt Measurement"    
    a+=1
    print("wrote titles")
    return a

def createtitles2(a):
    b=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_3w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V3W Measurement"
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W Measurement"   
    b+=1
    sheet.cell(row=a, column=b).value = "Freq [Hz]"
    b+=1
    sheet.cell(row=a, column=b).value = "Reference Voltage[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt X[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V]"
    b+=1
    sheet.cell(row=a, column=b).value = "V_1w_shunt Y[V] Standard dev."
    b+=1
    sheet.cell(row=a, column=b).value = "Sensitivity of V1W_shunt Measurement"    
    b+=1
    sheet.cell(row=a, column=b).value = "T of sample" 
    a+=1
    print("wrote titles")
    return a

def to_excel(measurements,c,d): #c=a d=b
    for a, i in enumerate(measurements):
        for b, elem in enumerate(i):
            sheet.cell(row=(a+c), column=(b+d)).value = elem
    book.save('hello.xlsx')
    return (c + len(measurements))

def setting():
    ask_lockin('RMOD?') #returns dynamic reserve mode (high reserve/normal/low noise)
    ask_lockin('OFLT?') #returns the time constant (10microsec-30kilosec)
    ask_lockin('OFSL?') #returns low pass filter slope (6/12/18/24 dB/oct) 
    ask_lockin('SYNC?') #returns synchronous filter (Off/On) 
    ask_lockin('IGND?') #returns input shield grounding (GROUND/FLOATING)
    ask_lockin('ICPL?') #returns input coupling (AC/DC)
    ask_lockin('ILIN?') #returns line notch filters (Out, Line in, 2xLine in, Both In)
##########################################################
#these functions are for scanning and returning data values
################################################################
'''
def stable(x_splits,stablity_cond): #change how stability is defined********************************W******
    x = x_splits 
    vg = int(len(x)/2)

    high_avg = x[:vg].mean()
    low_avg = x[vg:len(x)].mean()
    
    
    z = abs(high_avg - low_avg)
    if z < stablity_cond:
        return True
    else:
        return False
    '''
def stable(x_splits,stablity_cond): #change how stability is defined********************************W******
    
    
    maxi = np.amax(x_splits)
    mini = np.amin(x_splits)
    z = abs(maxi - mini)
    if z < stablity_cond:
        return True
    else:
        return False
           

def stabilize(lc_cmd,lc_val,ttt,stability_cond): #checks that the data is stable over 3rd line of code seconds
    tt=ttt
    tell_lockin(lc_cmd+str(lc_val))
    #run for double the time
    #use it anyway despite stability or throw data point away
    #make sure it can't go on infinite loop
    while 1:
        tell_lockin('REST') #clears and pauses buffer
        tell_lockin('STRT') #starts scan
        time.sleep(3)
        x = ask_lockin('TRCA?1,0,'+str(31*tt)) 
        x_split = x.split(',')
        x_split = x_split[:(len(x_split)-1)]
        x_split = np.array([float(i) for i in x_split])
    
        if (stable(x_split,stability_cond)):
            y = ask_lockin('TRCA?1,0,'+str(31*tt)) 
            y_split = y.split(',')
            y_split = y_split[:(len(y_split)-1)]
            y_split = np.array([float(i) for i in y_split])
            return x_split,y_split
            break
    
def calibrate(stb,max_stb_time,first_state,second_state,lc_cmd_state):
    t = max_stb_time
    
    tell_lockin('REST')
    
    tell_lockin(lc_cmd_state + str(first_state))
    tell_lockin('STRT')
    tell_lockin(lc_cmd_state + str(second_state))
    time.sleep(t)
    x = ask_lockin('TRCA?1,0,'+str(31*t))
    x = x[:len(x)-2] #change back to 2 if you get a syntax error
    x = x.split(',')
    x = [float(i) for i in x]
    x = np.array(x)
    spaced = t-3
    for i in range(spaced):
        avg = x[31*(i+1):31*(i+4)]
        if stable(avg,stb) == True:
            #print('')
            #print('stabilized after '+ str(i+1) + ' seconds')
            return i+1
            break
    #stb = float(input('reconsider stability condition as system is not stable after 15 seconds. Input new stability condition:  '))
    calibrate(stb,max_stb_time,first_state,second_state,lc_cmd_state)
    
def calibrate2(stb,max_stb_time,first_state,second_state,lc_cmd_state):
    t = max_stb_time
    print('')
    print('calibrating...')
    tell_lockin('REST')
    
    tell_lockin(lc_cmd_state + str(first_state))
    time.sleep(6)
    tell_lockin('STRT')
    tell_lockin(lc_cmd_state + str(second_state))
    time.sleep(t)
    x = ask_lockin('TRCA?1,0,'+str(31*t))
    x = x[:len(x)-2] #change back to 2 if you get a syntax error
    x = x.split(',')
    x = [float(i) for i in x]
    x = np.array(x)
    spaced = t-3
    for i in range(spaced):
        avg = x[31*(i+1):31*(i+4)]
        if stable(avg,stb) == True:
            print('')
            print('stabilized after '+ str(i+1) + ' seconds')
            return i+4
            break
    stb = float(input('reconsider stability condition as system is not stable after 15 seconds. Input new stability condition:  '))
    calibrate(stb,max_stb_time,first_state,second_state,lc_cmd_state)
        
    
  
def scan(harm,lc_list,tt,volt_q,val_f,stab_cond): #char is the command for what is being measured, listchar is the list containing the values that are to be scanned, stable_time is the amount of time spent at each value in listchar, tt is the amount of time taken into account for mean.x and stdv.x
    tell_lockin('HARM'+str(harm))
    output_x = []                                       #volt_q=True if scanning volts false if scanning freq
    stdv_x = []
    data = []
    
    #stab_cond = float(input("input stability condition(V):    "))
    
    
    if volt_q == 1:
        lc_list = listvolts
        lc_cmd = 'SLVL'
        na = 'V'
    else:
        lc_list = listfreq
        lc_cmd = 'FREQ'
        na = 'Hz'
    
    t = calibrate(stab_cond,15,lc_list,lc_cmd) #if shit is hitting the fan change back to 16
    
    print('scanning whole range')
    print('')
    sensitivity = ask_lockin('SENS?')
    tell_lockin('REST') #clears and pauses buffer
    tell_lockin('STRT') #starts scan
    
    for i in lc_list:
        tell_lockin(lc_cmd+str(i))
        time.sleep(t)
        
    x = ask_lockin('TRCA?1,0,'+str(31*len(lc_list)*t)) #outputs buffer
    x = x[:len(x)-1]
    print(x)
    y = ask_lockin('TRCA?2,0,'+str(31*len(lc_list)*t)) 
    y = y[:len(y)-1]
    
    for i,j in enumerate(lc_list):
        x_split = x.split(',')[((31*i*t)+(t-tt)*31):(31*(i+1)*t)]    #31 is number of measurements a second
        
        x_split = [float(i) for i in x_split]
        x_split = np.array(x_split)
        
        if stable(x_split,stab_cond) == False:
            print("not stable at: " + str(j)+na)
            x_split,y_split = stabilize(lc_cmd,j,tt,stab_cond) #make sure number of seconds spliced and number of seconds taken from stabilize are equal
        else: 
            print("stable at: " + str(j)+na)
            y_split = y.split(',')[((31*i*t)+(t-tt)):(31*(i+1)*t)]
            y_split = [float(i) for i in x_split]
            y_split = np.array(y_split)
        
        out_x = x_split.mean()
        out_y = y_split.mean()
        print('')
        print("output mean:  "+str(out_x))
        print('')
        sd_x = x_split.std()
        sd_y = y_split.std()
        
        #out_x,out_y,sd_x,sd_y= averaged_data(x_split,y_split)
        output_x.append(out_x)
        stdv_x.append(sd_x)
        
        
        if volt_q != 1:
            huh = val_f
            val_f = j
            j = huh
            
        data.append([val_f,j,out_x,sd_x,out_y,sd_y,sensitivity])
  
     
    return data,output_x,stdv_x

########################################################@##
#THESE functions calculate and plot values#################################
###########################################################

def calculate_plot1(v1_sh,r_sh,v1_):
    i_current = []
    i_cubbed = []
    r = []
    power = []
    ctr = 0
    for i in range(len(v1_)):
        for j in range(len(v1_[0])):
            print(type(v1_sh))
            i_current.append(v1_sh[i][j]/r_sh)     #i = v1w / r shunt
            i_cubbed.append(i_current[ctr]**3)            #i^3 
            power.append(v1_[i][j]*i_current[ctr])        #power = v1w * current
            r.append(v1_[i][j]/i_current[ctr])            #r = v1w / i
            ctr+=1
    return i_current,i_cubbed,power,r

def plot_ic_vs_v3(v3w,cub,stdv_v3w):
    v3w = np.array(v3w)
    stdv_v3w = np.array(stdv_v3w)
    cub = np.array(cub)
    plt.errorbar(cub,v3w.flatten(), stdv_v3w.flatten(),fmt='o') #0:4 -> 4-8
    plt.show()

def plot_r_vs_power(r,power):
    plt.errorbar(power,r,fmt='o') #0:4 -> 4-8
    plt.xlabel('Power[W]')
    plt.ylabel('R Sample[Ohm]')
    plt.show()
    
def plot_v3w_vs_ln2w(v3w,stdv_v3w,freqs):
    v3w = np.array(v3w)
    stdv_v3w = np.array(stdv_v3w)
    ln2w = [np.log(4*i*np.pi) for i in freqs]
    ln2w = ln2w + (list(reversed(ln2w))) +  ln2w + (list(reversed(ln2w)))
    plt.errorbar(v3w.flatten(),ln2w,stdv_v3w.flatten(),fmt='o') 
    plt.show()
 
#initialize variables
a = 1
b = 1
book = Workbook()
sheet = book.active    

    
# main
lock = serial.Serial('COM5', baudrate=19200, parity=serial.PARITY_NONE,
                         stopbits=serial.STOPBITS_ONE, bytesize=serial.EIGHTBITS, timeout=3)   

A = pd.read_excel (r'C:\Users\Pedro\Desktop\HelloWorld\SR830\sweep2.xlsx',sheet_name='Sheet1')
listfreq=A['freq']
listvolts=A['volts']


tell_lockin("OUTX0")
tell_lockin('SRAT9')

'''
listfreqrev = list(reversed(listfreq))

listvoltsrev = list(reversed(listvolts))

for i,k in zip(listvolts[0::1], listvolts[1::1]):
    print(i)
    print(k)
'''
listfreqrev = list(reversed(listfreq))
listvoltsrev = list(reversed(listvolts))
stb2 = 1e-7
y=[]
i=23
k=22
for i in range(10):
    y.append(calibrate(stb2,6,i,k,'FREQ'))
    



'''
harm = 1
while harm<3:
    harm+=2
    ctr = 1
    while ctr<3:
        y1=[]
        y2=[]
        y3=[]
        y4=[]
        
        ##########Scanning Frequencies############
        
        tell_lockin('HARM'+str(harm)) 
        if harm==1:
            tell_lockin('SENS26')
            stb2 = 0.0001
        if harm==3:
            tell_lockin('SENS22')
            stb2 = 249e-6
           
        tell_lockin('SLVL3.2')         ############change calibrates to calibrate2 and stb1 and stb2 to 0.00005
        
        
        t=time.time()
        tell_lockin('FREQ'+str(listfreq[0]))
        time.sleep(6)
        for i,k in zip(listfreq[0::1], listfreq[1::1]):
            y1.append(calibrate(stb2,6,i,k,'FREQ')) 
            print(y1)
            print('')
            
        tell_lockin('FREQ'+str(listfreq[len(listfreq)-1]))
        time.sleep(6)
        for i,k in zip(listfreqrev[0::1], listfreqrev[1::1]):
            y2.append(calibrate(stb2,6,i,k,'FREQ'))
            print(y2)
            print('')
            
        print('took: '+str(time.time()-t)+' to perform scan')
        print('')
        print('graph '+str(ctr)+' of frequencies with harmonic: '+str(harm))
        
        fig = plt.figure()
        plt.errorbar(listfreq[1:],y1,fmt='o') 
        plt.errorbar(listfreqrev[1:],y2,fmt='o') 
        fig.suptitle('Time to stabilize at different frequencies ('+str(harm)+'\u03C9)', fontsize=20)
        plt.xlabel('Frequency', fontsize=16)
        plt.ylabel('Time to stabilize', fontsize=16)
        plt.legend(['forward','reverse'])
        fig.savefig('calibrate_freq'+str(ctr)+'at harm'+str(harm)+'.jpg')
        plt.show()
        
        workbook = xlsxwriter.Workbook('calibrate_freq'+str(ctr)+'at harm'+str(harm)+'.xlsx') 
        worksheet = workbook.add_worksheet() 
        
        row = 0
        column = 0
        
        for item in listfreq:
            worksheet.write(row, column, item)
            row+=1
        
        column = 1
        row = 0
        
        for item in y1:
            worksheet.write(row, column, item)
            row+=1
        
        column = 2
        row = 0
        
        for item in listfreqrev:
            worksheet.write(row, column, item)
            row+=1
            
        column = 3
        row = 0 
        
        for item in y2:
            worksheet.write(row, column, item)
            row+=1    
            
        workbook.close()
        
        
        
        ##########Scanning Voltages#############
        
        tell_lockin('FREQ122')
        stb1 = stb2
        
        t = time.time()
        tell_lockin('SLVL'+str(listvolts[0]))
        time.sleep(6)
        for i,k in zip(listvolts[0::1], listvolts[1::1]):
            y3.append(calibrate(stb1,10,i,k,'SLVL'))
            
         
            
        
        tell_lockin('SLVL'+str(listvolts[len(listvolts)-1]))
        time.sleep(6)
        for i,k in zip(listvoltsrev[0::1], listvoltsrev[1::1]):
            y4.append(calibrate(stb1,10,i,k,'SLVL'))
            
        
        
        print('took: '+str(time.time()-t)+' to perform scan')
        print('')
        print('graph '+str(ctr)+' of voltages with harmonic: '+str(harm))
        fig = plt.figure()
        plt.errorbar(listvolts[1:],y3,fmt='o') 
        plt.errorbar(listvoltsrev[1:],y4,fmt='o') 
        fig.suptitle('Time to stabilize at different voltages ('+str(harm)+'\u03C9)', fontsize=20)
        plt.xlabel('Voltage', fontsize=16)
        plt.ylabel('Time to stabilize', fontsize=16)
        plt.legend(['forward','reverse'])
        fig.savefig('calibrate_volt'+str(ctr)+'at harm'+str(harm)+'.jpg')
        plt.show()
        
        workbook = xlsxwriter.Workbook('calibrate_volt'+str(ctr)+'at harm'+str(harm)+'.xlsx') 
        worksheet = workbook.add_worksheet() 
        
        row = 0
        column = 0
        
        for item in listvolts:
            worksheet.write(row, column, item)
            row+=1
        
        column = 1
        row = 0
        
        for item in y3:
            worksheet.write(row, column, item)
            row+=1
        
        column = 2
        row = 0
        
        for item in listvoltsrev:
            worksheet.write(row, column, item)
            row+=1
            
        column = 3
        row = 0    
        
        for item in y4:
            worksheet.write(row, column, item)
            row+=1    
            
        workbook.close()
        ctr+=1

lock.close()
'''